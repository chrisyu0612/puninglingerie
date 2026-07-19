"""
Import Full-Cup series from Excel into productdata/products/standard.
- One folder per series (Style No.)
- One image per colorway under colors/
- Never imports or writes Price
"""
from __future__ import annotations

import json
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

Image.MAX_IMAGE_PIXELS = None

XLSX = Path(r"d:\06 独立站相关\FB账号运营\A Full-Cup Series Product Table with Prices - English.xlsx")
SITE_ROOT = Path(__file__).resolve().parents[1]
STANDARD_ROOT = SITE_ROOT / "productdata" / "products" / "standard"
NEW_ROOT = SITE_ROOT / "productdata" / "products" / "new"
MAX_SIDE = 1400
JPEG_QUALITY = 82
ACCENTS = ["#4be683", "#6196ff", "#ffe36d", "#ff8a66", "#84e6ef", "#c6a7ff"]


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[\s_/]+", "-", text)
    text = re.sub(r"[^a-z0-9\-]+", "", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "color"


def style_slug(style_no: str) -> str:
    base = style_no.replace("\n", " ").strip()
    # Keep leading digits / code readable: style-201, style-222-full-cup-version
    slug = slugify(base)
    if not slug.startswith("style-"):
        slug = f"style-{slug}"
    return slug


def parse_styles():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["Sheet2"]
    styles = []
    pending = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        style = row[0]
        if style:
            style_text = str(style).replace("\n", " ").strip()
            style_text = re.sub(r"\s+", " ", style_text)
            pending = {
                "sheet_row": i,  # 1-indexed
                "image_row": i - 1,  # 0-indexed drawing row
                "style_no": style_text,
                "size": (str(row[1]).replace("\n", " ").strip() if row[1] else ""),
                "fabric": (str(row[15]).replace("\n", " ").strip() if row[15] else ""),
                "patent": bool(row[17]),
                # Price intentionally ignored: row[18]
                "colors": [],
            }
            styles.append(pending)
        elif pending is not None:
            names = []
            for col_idx in range(3, 15, 2):
                if col_idx < len(row) and row[col_idx]:
                    names.append(str(row[col_idx]).strip())
            pending["colors"] = names
            pending = None
    wb.close()
    return styles


def load_media_bytes():
    with zipfile.ZipFile(XLSX) as z:
        rels = z.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
        rid_to_media = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="../media/([^"]+)"', rels))
        drawing = z.read("xl/drawings/drawing1.xml").decode("utf-8")
        anchors = re.findall(
            r"<xdr:twoCellAnchor[\s\S]*?<xdr:from>\s*<xdr:col>(\d+)</xdr:col>[\s\S]*?<xdr:row>(\d+)</xdr:row>[\s\S]*?r:embed=\"(rId\d+)\"[\s\S]*?</xdr:twoCellAnchor>",
            drawing,
        )
        by_row: dict[int, list[tuple[int, str]]] = defaultdict(list)
        seen = set()
        for col, row, rid in anchors:
            media = rid_to_media.get(rid)
            if not media:
                continue
            key = (int(row), int(col), media)
            if key in seen:
                continue
            seen.add(key)
            by_row[int(row)].append((int(col), media))
        for row in by_row:
            by_row[row].sort(key=lambda x: x[0])

        media_bytes = {}
        for name in z.namelist():
            if name.startswith("xl/media/"):
                media_bytes[name.split("/")[-1]] = z.read(name)
        return by_row, media_bytes


def save_resized(raw: bytes, dest: Path):
    from io import BytesIO

    img = Image.open(BytesIO(raw))
    img = img.convert("RGB")
    w, h = img.size
    scale = min(1.0, MAX_SIDE / max(w, h))
    if scale < 1.0:
        img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    img.save(dest, format="JPEG", quality=JPEG_QUALITY, optimize=True)


def build_description(style: dict) -> str:
    parts = []
    if style["colors"]:
        parts.append("Colors: " + ", ".join(style["colors"]))
    if style["size"]:
        parts.append("Size: " + style["size"])
    if style["fabric"]:
        parts.append(style["fabric"])
    if style["patent"]:
        parts.append("Patent design.")
    return " | ".join(parts) if parts else "Full-Cup series style."


def clear_products():
    for root in (STANDARD_ROOT, NEW_ROOT):
        if root.exists():
            shutil.rmtree(root)
        root.mkdir(parents=True, exist_ok=True)


def pick_row_images(by_row: dict, row: int) -> list[tuple[int, str]]:
    images = by_row.get(row, [])
    # Prefer color slots: D/F/H/... (cols 3,5,7,...)
    slotted = [(c, m) for c, m in images if c >= 3 and c % 2 == 1]
    source = slotted or images
    ordered = []
    used_media = set()
    for col, media in source:
        if media in used_media:
            continue
        used_media.add(media)
        ordered.append((col, media))
    return ordered


def assign_image_rows(styles: list, by_row: dict) -> dict[str, list[tuple[int, str]]]:
    """Assign image rows without stealing another series' primary row."""
    reserved = {s["image_row"] for s in styles}
    assignment: dict[str, list[tuple[int, str]]] = {}
    used_rows: set[int] = set()

    # Pass 1: exact primary row
    pending = []
    for style in styles:
        key = style["style_no"]
        images = pick_row_images(by_row, style["image_row"])
        if images:
            assignment[key] = images
            used_rows.add(style["image_row"])
        else:
            pending.append(style)

    # Pass 2: only unassigned non-reserved rows near the style
    for style in pending:
        key = style["style_no"]
        idx = styles.index(style)
        lo = styles[idx - 1]["image_row"] + 1 if idx > 0 else 0
        hi = styles[idx + 1]["image_row"] - 1 if idx + 1 < len(styles) else style["image_row"] + 3
        found = []
        for row in range(lo, hi + 1):
            if row in used_rows or row in reserved:
                continue
            images = pick_row_images(by_row, row)
            if images:
                found = images
                used_rows.add(row)
                break
        assignment[key] = found

    return assignment


def import_all():
    if not XLSX.exists():
        raise SystemExit(f"Excel not found: {XLSX}")

    styles = parse_styles()
    by_row, media_bytes = load_media_bytes()
    clear_products()
    row_map = assign_image_rows(styles, by_row)

    print(f"styles={len(styles)} image_rows={len(by_row)}")
    imported = 0
    for index, style in enumerate(styles):
        slug = style_slug(style["style_no"])
        folder = STANDARD_ROOT / slug
        colors_dir = folder / "colors"
        colors_dir.mkdir(parents=True, exist_ok=True)

        ordered = row_map.get(style["style_no"], [])
        # Cap to declared color count when Excel has extra decorative images in the slots
        color_names = style["colors"] or []
        if color_names and len(ordered) > len(color_names):
            ordered = ordered[: len(color_names)]

        color_entries = []
        for i, (col, media) in enumerate(ordered):
            name = color_names[i] if i < len(color_names) else f"Color {i + 1}"
            color_slug = slugify(name)
            dest_name = f"{color_slug}.jpg"
            n = 2
            while (colors_dir / dest_name).exists():
                dest_name = f"{color_slug}-{n}.jpg"
                n += 1
            dest = colors_dir / dest_name
            raw = media_bytes.get(media)
            if not raw:
                print(f"  missing media {media} for {slug}")
                continue
            save_resized(raw, dest)
            rel = f"productdata/products/standard/{slug}/colors/{dest_name}"
            color_entries.append({"id": Path(dest_name).stem, "name": name, "image": rel})

        if not color_entries:
            print(f"  SKIP {slug}: no images")
            shutil.rmtree(folder, ignore_errors=True)
            continue

        src_cover = SITE_ROOT / color_entries[0]["image"]
        shutil.copyfile(src_cover, folder / "cover.jpg")

        tags = ["fullcup", "bra"]
        if style["patent"]:
            tags.append("patent")

        info = {
            "id": slug,
            "title": style["style_no"],
            "subtitle": "Full-Cup Series",
            "description": build_description({**style, "colors": [c["name"] for c in color_entries]}),
            "size": style["size"],
            "fabric": style["fabric"],
            "patent": style["patent"],
            "colors": color_entries,
            "tags": tags,
            "accent": ACCENTS[index % len(ACCENTS)],
            "tilt": f"{(-2 + (index % 5))}deg",
            "linkText": "询盘此系列",
            "sort": (index + 1) * 10,
            "featured": index < 3,
        }
        (folder / "info.json").write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
        imported += 1
        print(f"OK {slug} colors={len(color_entries)}/{len(color_names) or len(color_entries)}")

    print(f"Imported {imported} series into {STANDARD_ROOT}")


if __name__ == "__main__":
    import_all()
